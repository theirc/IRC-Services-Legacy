import json
import logging

import re
from admin_panel.utils import push_service_to_transifex, pull_completed_service_from_transifex, \
    get_service_transifex_info
from api.utils import generate_translated_fields
from api.v2 import serializers as serializers_v2
from api.v2.serializers import ServiceImageSerializer, ServiceAreaSerializer, CreateProviderSerializer, ProviderSerializer, ProviderTypeSerializer
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.contrib.gis.geos import Point, Polygon
from django.contrib.gis.measure import D
from django.db.models import Q
from django.core.exceptions import ValidationError as DjangoValidationError
from django.db.transaction import atomic
from django.db.utils import IntegrityError
from regions.models import GeographicRegion
from rest_framework import permissions, parsers
from rest_framework import status, viewsets, mixins
from rest_framework.decorators import list_route, detail_route
from rest_framework.exceptions import ValidationError as DRFValidationError
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from services.models import Service, Provider, ServiceType, ServiceArea, ServiceTag, ProviderType, ServiceConfirmationLog, ContactInformation
from .utils import StandardResultsSetPagination, FilterByRegionMixin
from ..filters import ServiceFilter, CustomServiceFilter, RelativesServiceFilter, WithParentsServiceFilter, PrivateServiceFilter
from django_filters import rest_framework as django_filters
from django.db.models import Count
from rest_framework.authtoken.models import Token
from django.contrib.sessions.models import Session
from django.utils import timezone

logger = logging.getLogger(__name__)
import openpyxl
import base64
from io import BytesIO
import tempfile
from django.conf import settings

from rest_framework import filters

from haystack.query import SearchQuerySet
from django.db.models import Case, When


class ServiceAreaViewSet(viewsets.ModelViewSet):
    permission_classes = [AllowAny]
    queryset = ServiceArea.objects.all()
    serializer_class = ServiceAreaSerializer


class SearchFilter(filters.SearchFilter):
    def filter_queryset(self, request, queryset, view):
        if 'service-management' in request.get_full_path():
            return super(SearchFilter, self).filter_queryset(request, queryset, view)

        params = request.query_params.get(self.search_param, '')
        if not params:
            return queryset

        pk_list = [o.pk for o in SearchQuerySet().filter(content=params)]
        preserved = Case(*[When(pk=pk, then=pos)
                           for pos, pk in enumerate(pk_list)])
        return queryset.filter(pk__in=pk_list).order_by(preserved)


class ProviderViewSet(FilterByRegionMixin, viewsets.ModelViewSet):
    queryset = Provider.objects.all()
    serializer_class = serializers_v2.ProviderSerializer
    pagination_class = StandardResultsSetPagination

    # All the text fields that are used for full-text searches (?search=XXXXX)
    search_fields = generate_translated_fields('name', False) \
        + generate_translated_fields('description', False) \
        + generate_translated_fields('focal_point_name', False) \
        + generate_translated_fields('address', False) \
        + generate_translated_fields('type__name', False) \
        + ['phone_number'] \
        + ['website'] \
        + ['number_of_monthly_beneficiaries'] \
        + ['focal_point_phone_number']

    def update(self, request, *args, **kwargs):
        """On change to provider via the API, notify via JIRA"""
        response = super().update(request, *args, **kwargs)
        provider = Provider.objects.get(id=request.data["id"])
        if request.data["is_frozen"]:
            """ Delete all tokens and sessions for users linked to provider """
            users = [user for user in list(
                [provider.user]) + list(provider.team.all()) if user]
            Token.objects.filter(user__in=users).delete()
            for user in users:
                user_sessions = []
                all_sessions = Session.objects.filter(
                    expire_date__gte=timezone.now())
                for session in all_sessions:
                    session_data = session.get_decoded()
                    if str(user.pk) == session_data.get('_auth_user_id'):
                        user_sessions.append(session.pk)
                Session.objects.filter(pk__in=user_sessions).delete()
        return response

    @list_route(methods=['get'], permission_classes=[IsAuthenticated])
    def my_providers(self, request):
        filtered = self.filter_queryset(self.get_queryset())
        my = filtered.filter(user=request.user) | request.user.providers.all()

        return Response([self.get_serializer_class()(a, context={'request': request}).data for a in my])

    @list_route(methods=['post'], permission_classes=[AllowAny])
    def create_provider(self, request, *args, **kwargs):
        """
        Customized "create provider" API call.

        This is distinct from the built-in 'POST to the list URL'
        call because we need it to work for users who are not
        authenticated (otherwise, they can't register).

        Expected data is basically the same as for creating a provider,
        except that in place of the 'user' field, there should be an
        'email' and 'password' field.  They'll be used to create a new user,
        send them an activation email, and create a provider using
        that user.
        """
        with atomic():  # If we throw an exception anywhere in here, rollback all changes
            serializer = CreateProviderSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)

            # Create User
            user = get_user_model().objects.create_user(
                email=request.data['email'],
                password=request.data['password'],
                is_active=False
            )
            provider_group, _ = Group.objects.get_or_create(name='Providers')
            user.groups.add(provider_group)

            # Create Provider
            data = dict(request.data, user=user.get_api_url())
            serializer = ProviderSerializer(
                data=data, context={'request': request})
            serializer.is_valid(raise_exception=True)
            serializer.save()  # returns provider if we need it
            headers = self.get_success_headers(serializer.data)

            # If we got here without blowing up, send the user's activation email
            user.send_activation_email(
                request.site, request, data['base_activation_link'])
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    @list_route(methods=['post'], permission_classes=[AllowAny])
    def claim_service(self, request, *args, **kwargs):
        try:
            with atomic():
                if request.user.groups.filter(name='Providers').exists():
                    user = request.user
                else:
                    try:
                        user = get_user_model().objects.create_user(
                            email=request.data['email'],
                            password=get_user_model().objects.make_random_password(),
                            is_active=False
                        )
                    except DjangoValidationError:
                        raise DjangoValidationError(
                            'User with this email exists')
                    provider_group, _ = Group.objects.get_or_create(
                        name='Providers')
                    user.groups.add(provider_group)

                # Create or update Provider
                data = dict(request.data, user=user.get_api_url())
                if hasattr(user, 'provider'):
                    serializer = ProviderSerializer(
                        user.provider, data=data, context={'request': request})
                else:
                    serializer = ProviderSerializer(
                        data=data, context={'request': request})
                try:
                    serializer.is_valid(raise_exception=True)
                except DRFValidationError:
                    raise DjangoValidationError(
                        'All fields needs to be filled.')
                serializer.save()
                headers = self.get_success_headers(serializer.data)
                service = Service.objects.get(id=request.data['service_id'])
                user.send_activation_email_to_staff(
                    request, service, serializer.instance)
                return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
        except DjangoValidationError as ex:
            return Response('; '.join(ex.messages), status=status.HTTP_400_BAD_REQUEST)
        except Exception as ex:
            logger.error(
                'There was an error during claim',
                exc_info=True,
            )
            return Response(str(ex), status=status.HTTP_400_BAD_REQUEST)

    @detail_route(methods=['get'], permission_classes=[permissions.DjangoObjectPermissions])
    def export_services(self, request, pk=None, *args, **kwargs):
        obj = self.get_queryset().filter(pk=pk).first()
        services = obj.services.all()

        book = openpyxl.Workbook()
        sheet = book.get_active_sheet()

        provider_services = []

        for s in services:
            s = serializers_v2.ServiceExcelSerializer(s).data

            provider_services.append(s)

        headers = list(serializers_v2.ServiceExcelSerializer.FIELD_MAP.keys())
        human_headers = list(
            serializers_v2.ServiceExcelSerializer.FIELD_MAP.values())

        for col in range(0, len(human_headers)):
            sheet.cell(column=col + 1, row=1).value = human_headers[col]

        for row in range(0, len(provider_services)):
            for col in range(0, len(headers)):
                sheet.cell(column=col + 1, row=row +
                           2).value = provider_services[row][headers[col]]

        book_data = BytesIO()
        book.save(book_data)
        book_data.seek(0)

        return Response({
            "data": base64.b64encode(book_data.read()),
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }, content_type="application/json")

    @detail_route(methods=['GET'])
    def impersonate_provider(self, request, pk):
        request.session['selected-provider'] = pk

        return Response({})
    
    @detail_route(methods=['GET'])
    def stop_impersonate_provider(self, request, pk):
        if 'selected-provider' in request.session:
            del request.session['selected-provider']
            request.session.modified = True
            logger.error(
                    '**** STOP IMPERSONATING'
                )

        return Response({})

    @detail_route(methods=['post'],
                  permission_classes=[permissions.DjangoObjectPermissions],
                  parser_classes=[parsers.MultiPartParser])
    def import_services(self, request, pk=None, *args, **kwargs):
        obj = self.get_queryset().filter(pk=pk).first()
        file = request.data['file']

        with tempfile.NamedTemporaryFile(suffix='.xlsx') as nf:
            nf.delete = True
            with open(nf.name, 'w+') as fp:
                fp.write(str(file.read()))

            book = openpyxl.load_workbook(file)
        sheet = book.get_active_sheet()

        reversed_field_map = dict(
            [(d[1], d[0]) for d in serializers_v2.ServiceExcelSerializer.FIELD_MAP.items()])

        headers = [sheet.cell(row=1, column=c).value for c in range(
            1, sheet.max_column + 1)]
        headers = [reversed_field_map[h]
                   if h in reversed_field_map else h for h in headers]

        rows = [[sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
                for r in range(2, sheet.max_row + 1)]

        rows_to_use = [dict(zip(headers, r)) for r in rows]
        errors = {}

        try:
            with atomic() as t:
                for a in rows_to_use:
                    for f in ["name_{}".format(k) for k, v in settings.LANGUAGES]:
                        if f not in a or not a[f]:
                            a[f] = ''
                    for f in ["description_{}".format(k) for k, v in settings.LANGUAGES]:
                        if f not in a or not a[f]:
                            a[f] = ''
                    for f in ["address_{}".format(k) for k, v in settings.LANGUAGES]:
                        if f not in a or not a[f]:
                            a[f] = ''

                    serialized = serializers_v2.ServiceExcelSerializer(
                        data=a, partial=True)
                    if serialized.is_valid():
                        if 'delete' in a and a['delete']:
                            service = Service.objects.get(id=a['id'])
                            service.delete()
                            continue

                        if a['id']:
                            service = Service.objects.get(id=a['id'])
                        else:
                            region = a['region']
                            region = GeographicRegion.objects.filter(
                                id=region).first()
                            serialized.validated_data.pop('region', '')
                            service = Service(
                                region=region, provider=obj, **serialized.validated_data)
                            service.save()

                        if 'location' in a and a['location']:
                            location = Point(
                                *reversed([float(b.strip()) for b in a['location'].split(',')]), srid=4326)
                        else:
                            location = service.location

                        if 'location' in serialized.validated_data:
                            del serialized.validated_data['location']

                        Service.objects.filter(id=service.id).update(
                            location=location, **serialized.validated_data)
                    else:
                        errors = serialized.errors
                        raise IntegrityError

                return Response(None, status=204)
        except IntegrityError:
            return Response(errors, status=400)


class PrivateProviderViewSet(FilterByRegionMixin, viewsets.ModelViewSet):
    queryset = Provider.objects.all()
    serializer_class = serializers_v2.ProviderSerializer
    pagination_class = StandardResultsSetPagination


class PrivateServiceViewSet(FilterByRegionMixin, viewsets.ModelViewSet):
    filter_class = PrivateServiceFilter
    queryset = Service.objects.select_related(
        'provider',
        'type',
        'region',
    ).prefetch_related('selection_criteria', 'tags', 'types', 'contact_information').all()

    serializer_class = serializers_v2.ServiceSerializer
    pagination_class = StandardResultsSetPagination
    filter_backends = (django_filters.DjangoFilterBackend,
                       filters.OrderingFilter, SearchFilter)

    def get_queryset(self):
        qs = super(PrivateServiceViewSet, self).get_queryset()
        if not (hasattr(self.request, 'user') and self.request.user.is_superuser):
            providers = self.request.user.all_providers
            qs = qs.filter(Q(status__in=[Service.STATUS_CURRENT]) | Q(
                status__in=[Service.STATUS_PRIVATE], provider__in=providers))
        return qs


class ServiceViewSet(viewsets.ModelViewSet):
    filter_class = ServiceFilter
    is_search = False
    is_preview = False

    # The queryset is only here so DRF knows the base model for this View.
    # We override it below in all cases.
    queryset = Service.objects.select_related(
        'provider',
        'type',
        'region',
    ).prefetch_related('selection_criteria', 'tags', 'types', 'contact_information').all()

    serializer_class = serializers_v2.ServiceSerializer
    pagination_class = StandardResultsSetPagination
    search_fields = ()
    filter_backends = (django_filters.DjangoFilterBackend,
                       filters.OrderingFilter, SearchFilter)

    def get_search_fields(self):
        if 'service-management' in self.request.get_full_path():
            return generate_translated_fields('name', False) \
                + generate_translated_fields('type__name', False) \
                + ['region__name'] \
                + ['status']
        else:
            return generate_translated_fields('additional_info', False) \
                + ['cost_of_service'] \
                + generate_translated_fields('description', False) \
                + generate_translated_fields('name', False) \
                + generate_translated_fields('type__comments', False) \
                + generate_translated_fields('type__name', False) \
                + generate_translated_fields('provider__description', False) \
                + generate_translated_fields('provider__focal_point_name', False) \
                + ['provider__focal_point_phone_number'] \
                + generate_translated_fields('provider__address', False) \
                + generate_translated_fields('provider__name', False) \
                + generate_translated_fields('provider__type__name', False) \
                + ['provider__phone_number', 'provider__website', 'provider__user__email'] \
                + generate_translated_fields('selection_criteria__text', False) \
                + ['region__slug', 'tags__name']

    def get_queryset(self):
        # Get filter type
        if self.request.query_params.get('filter') == 'wide':
            self.filter_class = CustomServiceFilter
        elif self.request.query_params.get('filter') == 'relatives':
            self.filter_class = RelativesServiceFilter
        elif self.request.query_params.get('filter') == 'with-parents':
            self.filter_class = WithParentsServiceFilter
        else:
            self.filter_class = ServiceFilter

        # Only make visible the Services owned by the current provider
        if self.is_search:
            qs = self.queryset.filter(status=Service.STATUS_CURRENT).order_by(
                'location', 'provider__name_en')
        elif self.is_preview:
            qs = self.queryset.filter(
                status__in=[Service.STATUS_CURRENT, Service.STATUS_DRAFT])
        else:
            qs = super().get_queryset()
        self.search_fields = self.get_search_fields()

        if self.request.query_params.get('near'):
            near_query = self.request.query_params.get('near')
            near_km = float(self.request.query_params.get('near_km', '5'))
            point = Point([float(a) for a in near_query.split(',')])

            qs = qs.filter(location__distance_lte=(point, D(km=near_km)))

        if self.request.query_params.get('bounds'):
            bounds_query = self.request.query_params.get('bounds')
            if ';' in bounds_query:
                bounds = [float(a) for x in [b.split(',')
                                             for b in bounds_query.split(';')] for a in x]
            else:
                bounds = [float(a) for a in bounds_query.split(',')]

            if len(bounds) == 4:
                west, north, east, south = bounds
                poly = Polygon.from_bbox((west, north, east, south))
                qs = qs.filter(location__within=poly)

        return qs

    @list_route(methods=['get'], permission_classes=[AllowAny])
    def search(self, request, *args, **kwargs):
        """
        Public API for searching public information about the current services
        """
        self.is_search = True
        #self.serializer_class = serializers_v2.ServiceSearchSerializer
        self.serializer_class = serializers_v2.ServiceSearchResultListSerializer
        return super().list(request, *args, **kwargs)
    
    @list_route(methods=['get'], permission_classes=[AllowAny])
    def searchlist(self, request, *args, **kwargs):
        
        self.is_search = True    self.serializer_class = serializers_v2.ServiceSearchResultListSerializer
        return super().list(request, *args, **kwargs)

    @list_route(methods=['get'], permission_classes=[AllowAny])
    def preview(self, request, *args, **kwargs):
        """
        Public API for previewing services (for statuses: draft or current)
        """
        self.is_preview = True
        self.serializer_class = serializers_v2.ServiceSearchSerializer
        return super().list(request, *args, **kwargs)

    def get_serializer_class(self):
        ob = self
        action = getattr(self, 'action')
        if getattr(self, 'action') == 'create':
            return serializers_v2.ServiceCreateSerializer        
        if getattr(self, 'action') == 'searchlist':
            return serializers_v2.ServiceSearchResultListSerializer    
        if getattr(self, 'action') == 'search':
            return serializers_v2.ServiceSearchResultListSerializer    
        if 'service-management' in self.request.get_full_path():
            return serializers_v2.ServiceManagementSerializer
        return serializers_v2.ServiceSerializer

    @detail_route(methods=['post'])
    def push_service_to_transifex(self, request, **kwargs):
        result = push_service_to_transifex(kwargs['pk'])
        return Response(result.info)

    @detail_route(methods=['get'])
    def pull_service_from_transifex(self, request, **kwargs):
        pulled_languages = pull_completed_service_from_transifex(kwargs['pk'])
        return Response(pulled_languages)

    @detail_route(methods=['get'])
    def get_service_transifex_data(self, request, **kwargs):
        r = get_service_transifex_info(kwargs['pk'])
        try:
            transifex_text = json.loads(r.text)
            languages = [i[0] for i in settings.LANGUAGES_CMS]
            output = dict((k, v['completed'])
                          for k, v in transifex_text.items() if k in languages)
        except ValueError:
            if r.status_code == 404:
                output = {
                    'errors': 'Service has not been sent to Transifex yet.'
                }
            else:
                output = {
                    'errors': 'An error occurred from Transifex (%s).' % r.text
                }

        return Response(output)

    def partial_update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', True)
        instance = self.get_object()
        serializer = ServiceImageSerializer(
            instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response({'status': 204, 'image': settings.MEDIA_URL + str(instance.image.name)})

    @detail_route(methods=['post'])
    def duplicate(self, request, **kwargs):
        service_id = kwargs.get('pk')
        new_name = request.query_params.get('new_name')
        if service_id and new_name:
            service_to_copy = Service.objects.prefetch_related(
                'tags', 'types').get(id=service_id)
            tags = service_to_copy.tags.all()
            types = service_to_copy.types.all()
            contact_information = service_to_copy.contact_information.all()
            # Clear service data
            service_to_copy.pk = None
            service_to_copy.slug = None
            service_to_copy.name = new_name
            service_to_copy.status = Service.STATUS_DRAFT
            service_to_copy.save()
            # Fill all fragile data
            new_name = re.sub('[\W-]+', '-', new_name.lower())
            new_slug = service_to_copy.region.slug + '_' + \
                str(service_to_copy.provider.id) + '_' + new_name
            services = Service.objects.filter(slug=new_slug)
            if services:
                new_slug = new_slug + '_' + str(service_to_copy.id)
            service_to_copy.slug = new_slug
            service_to_copy.tags.add(*tags)
            service_to_copy.types.add(*types)
            service_to_copy.contact_information.add(*contact_information)
            service_to_copy.save()
            return Response({'service_id': service_to_copy.id}, status=201)
        else:
            return Response({'error': 'Missing service id or service name'}, status=400)

    @detail_route(methods=['post'])
    def archive(self, request, **kwargs):
        service_id = kwargs.get('pk')
        if service_id:
            Service.objects.filter(id=service_id).update(
                status=Service.STATUS_ARCHIVED)
            return Response(None, status=200)
        else:
            return Response({'error': 'Missing service id'}, status=400)

    @detail_route(methods=['get'], permission_classes=[AllowAny])
    def get_same_coordinates_services(self, request, **kwargs):
        requested_service_id = kwargs.get('pk')
        filtered_by_coordinates = []
        max_distance_m = 5
        if requested_service_id:
            requested_service = Service.objects.get(id=requested_service_id)
            if requested_service.location:
                filtered_by_region = self.filter_queryset(self.get_queryset())
                filtered_by_coordinates = filtered_by_region.filter(
                    location__distance_lt=(
                        requested_service.location, D(m=max_distance_m)),
                    status=Service.STATUS_CURRENT) \
                    .exclude(id=requested_service_id)
        else:
            return Response({'error': 'Missing service id'}, status=400)

        serializer = serializers_v2.ServiceSerializer(
            filtered_by_coordinates, many=True, context={'request': request})
        return Response({'results': serializer.data}, status=200)


class ServiceTypeViewSet(viewsets.ModelViewSet):
    """
    Look up service types.
    """
    permission_classes = [permissions.DjangoModelPermissionsOrAnonReadOnly]
    queryset = ServiceType.objects.order_by('number')
    serializer_class = serializers_v2.ServiceTypeSerializer
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        print('GET', self.request.GET, self.request.META['HTTP_ACCEPT_LANGUAGE'])
        if 'region' in self.request.GET:
            region = self.request.GET['region']
            return self.queryset.filter(
                (
                    Q(service__region__slug=region) |
                    Q(service__region__parent__slug=region) |
                    Q(service__region__parent__parent__slug=region)
                ) & Q(service__status=Service.STATUS_CURRENT)
            ).annotate(service_count=Count('service')).filter(service_count__gt=0).distinct().order_by('number')
        else:
            return ServiceType.objects.all().order_by('number')


class CustomServiceTypeViewSet(viewsets.ModelViewSet):
    """
    Look up service types. Returns only those service types if there are services with this type
    (if GeographicRegion slug will be provided in request then results will be for all levels related to that region).
    """
    permission_classes = [permissions.DjangoModelPermissionsOrAnonReadOnly]
    queryset = Service.objects.prefetch_related('types'). \
        filter(types__isnull=False, status=Service.STATUS_CURRENT). \
        distinct('id')
    serializer_class = serializers_v2.CustomServiceTypeSerializer
    pagination_class = StandardResultsSetPagination

    filter_class = RelativesServiceFilter

    @list_route(methods=['GET'])
    def used_types(self, request, *args, **kwargs):
        filtered = self.filter_queryset(self.get_queryset())
        distincted = filtered.values_list('types__id', flat=True).distinct()
        types = ServiceType.objects.filter(id__in=distincted)
        return Response([serializers_v2.ServiceTypeSerializer(t, context={'request': request}).data for t in types])


class ProviderTypeViewSet(viewsets.ModelViewSet):
    """
    Look up provider types.

    """
    permission_classes = [permissions.DjangoModelPermissionsOrAnonReadOnly]
    queryset = ProviderType.objects.all()
    serializer_class = ProviderTypeSerializer
    pagination_class = StandardResultsSetPagination


class ServiceTagViewSet(viewsets.ModelViewSet):
    """
    Look up service tags with creating posibility
    """
    queryset = ServiceTag.objects.all()
    serializer_class = serializers_v2.ServiceTagSerializer


class ConfirmationLogsViewSet(viewsets.ModelViewSet):
    """
    Look up service Confirmation Logs
    """
    filter_class = ServiceFilter
    queryset = Service.objects.prefetch_related('confirmation_logs').all()
    serializer_class = serializers_v2.ServiceConfirmationLogsSerializer
    pagination_class = StandardResultsSetPagination


class ConfirmationLogListViewSet(viewsets.ModelViewSet):
    """
    Look up for all (last) Confirmation Logs
    """
    queryset = ServiceConfirmationLog.objects.order_by(
        'service', '-date').distinct('service')
    serializer_class = serializers_v2.ServiceConfirmationLogListSerializer
    pagination_class = StandardResultsSetPagination
